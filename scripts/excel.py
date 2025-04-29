import os
import sys
import yaml
import re
import jsonref
import logging
import markdown
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor, OneCellAnchor


COLUMN_DEFS = """
property:
    title: Technical Property
    width: 30
    style: bold
    description: |
        The technical name of the data attribute, as specified in the PACT Tech Specs 3.0. 
        The Data Model of the specification contains properties grouped in data types.
attribute:
    title: Methodology Attribute
    width: 20
    description: |
        The name of the data attribute as specified in the PACT Methodology v3.0.
        Note that not all attributes are listed in the Methodology, as some fields are purely technical.
section:
    title: Methodology Section
    width: 20
    style: small
    description: |
        A reference to the relevant section in the PACT Methodology v3.0.
reporting:
    title: Reporting Rule
    width: 15
    description: $$
description:
    title: Description
    width: 60
    style: text
    description: |
        The description of the data attribute, as specified in the PACT Tech Specs 3.0.
category:
    title: Category
    width: 17
    style: small
    description: |
        The category of the data attribute.
unit:
    title: Unit
    width: 17
    style: small
    description: |
        The unit of the data of the given attribute (e.g. kgCO2e / declaredUnit)
comment:
    title: Comment
    width: 30
    style: small
    description: |
        A comment on the data attribute. This may include information on how to calculate the value, or other relevant information.
accepted:
    title: Accepted Value(s)
    width: 20
    style: text
    description: |
        A description of the accepted values of the data attribute - some fields are limited to only a certain set of "valid values", whereas other fields can be free form text, booleans (true/false) or numbers.
example1:
    title: Example 1
    width: 30
    style: text
    description: |
        An example of how the data attribute could be populated (note exact syntax of each field is dependent on the solution used for data entry)
example2:
    title: Example 2
    width: 30
    style: text
    description: |
        An example of how the data attribute could be populated (note exact syntax of each field is dependent on the solution used for data entry)
example3:
    title: Example 3
    width: 30
    style: text
    description: |
        An example of how the data attribute could be populated (note exact syntax of each field is dependent on the solution used for data entry)
#mandatory:
#    title: Mandatory
#    description: >
#        An indication whether the field is mandatory or not. M indicates mandatory, O indicates optional, and M2025 indicates optional until 2025, when the field becomes mandatory.
"""

COLUMNS = yaml.safe_load(COLUMN_DEFS)

filter_text = {
    "See [[#lifecycle]] for details.": "",
    "([[!RFC8141|URN]])": "(RFC8141)",
    "See [[#validity-period]] for more details.": "",
    "See <{DataModelExtension}> for details.": "",
    #"See [=PACT Methodology=] for details.": "",
    #"(See [=PACT Methodology=])": "",
    #"See [=PACT Methodology=].": ""
}


class ExcelWriter:
    def __init__(self, output_path: str):
        self.output_path = output_path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "PACT Simplified Data Model"
        self.ws.sheet_view.zoomScale = 140

        # Define the columns of the worksheet
        self.columns = COLUMNS
        i = 0
        for column in self.columns.values():
            column['index'] = chr(ord('A') + i)
            column['index-nr'] = i
            i += 1
        # Define the styles for the worksheet
        fontname           = "Aptos Narrow"
        color_title        = "08094C"
        color_table_header = "2A4879" 
        color_header       = "489F81"
        self.styles = dict(
            title = dict(font = fontname, size = 16, bold = False, bgcolor = color_title, fgcolor = "FFFFFF"),
            subtitle = dict(font = fontname, size = 16, bold = False, bgcolor = color_title, fgcolor = "FFFFFF"),
            header = dict(font = fontname, bold = True, bgcolor = color_table_header, fgcolor = "FFFFFF"),
            header_desc = dict(font = fontname, size = 10, bgcolor = color_table_header, fgcolor = "FFFFFF"),
            subheader = dict(font = fontname, bold = True, bgcolor = color_header, fgcolor = "FFFFFF"),
            normal = dict(font = fontname), # bgcolor="EEECE2", border=True)
            bold = dict(font = fontname, bold = True),
            obsolete = dict(font = fontname, strike = True, fgcolor = "95261F"),
            small = dict(font = fontname, size = 9, fgcolor = "606060")
        )

        # Set cell widths
        for column in self.columns.values():
            self.ws.column_dimensions[column["index"]].width = column.get("width", 15)

    def save(self):
        # Apply word-wrap alignment for columns description and examples
        for column in self.columns.values():
            for cell in self.ws[column["index"]]:
                cell.alignment = Alignment(
                    indent = cell.alignment.indent, 
                    wrap_text=column.get("style", "") == "text",
                    vertical="top"
                    )
        # wrap all header descriptions
        for cell in self.ws[4]:
            cell.alignment = Alignment(
                indent = cell.alignment.indent, 
                wrap_text=column.get("style", "") == "text",
                vertical="top"
                )
        # HACK: merge reporting rules and description cells
        self.ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)

        # HACK: force style on data model extensions
        self.ws.cell(self.ws.max_row-3, 1).value = "extensions: DataModelExtension[]"
        self.ws.cell(self.ws.max_row-3, 2).value = "Data model extensions"
        self.format(self.ws[self.ws.max_row-3], self.styles['subheader'])

        # Insert the PACT logo
        img = Image("./assets/logo-dark-margin.png")
        img.width = img.width / 5
        img.height = img.height / 5
        self.ws.add_image(img, "A1")
        self.ws.row_dimensions[1].height = img.height * 4 / 3
        self.ws["A1"].alignment = Alignment(vertical="bottom")

        # Save the file
        self.wb.save(self.output_path)
        
    def write_title(self, title):
        row = ["" for column in self.columns.values()]
        row[0] = title
        self.ws.append(row)
        self.format(self.ws[1], self.styles['title'])

    def write_subtitle(self, subtitle):
        row = ["" for column in self.columns.values()]
        row[0] = subtitle
        self.ws.append(row)
        self.format(self.ws[self.ws.max_row], self.styles['subtitle'])

    def write_header(self):
        self.ws.append(column.get("title","") for column in self.columns.values())
        self.format(self.ws[self.ws.max_row], self.styles['header'])
        self.ws.append(column.get("description","") for column in self.columns.values())
        self.format(self.ws[self.ws.max_row], self.styles['header_desc'])    

    def write_type(self, items: list):
        self.ws.append(items)
        self.format(self.ws[self.ws.max_row], self.styles['subheader'])

    def write_property(self, object: dict):
        row = [str(object[name]) for name in self.columns.keys()]
        # Append the row to the worksheet
        self.ws.append(row)

        # Apply styles to the row, first col (property name) will be bold
        self.format(self.ws[self.ws.max_row], self.styles["normal"])
        for column in self.columns.values():        
            print(column)
            style = column.get("style", "normal")
            if style != "normal" and style != "text":
                self.format(self.ws[self.ws.max_row][column["index-nr"]], self.styles[style])
        # Strike-through obsolete or deprecated properties
        # if info.get("deprecated") or info.get("obsolete"):
        #    format(ws[ws.max_row], obsolete_style)
        # Indent the first cell of the row just added
        # ws[ws.max_row][0].alignment = Alignment(indent=level)

    
    def format(self, item, style):
        alignment = None

        #if style.get("word_wrap"):
        #    alignment = Alignment(vertical="top", wrap_text=style.get("word_wrap"))
        #else: 
        # if style.get("vertical_align"):
        #     alignment = Alignment(vertical=style.get("vertical_align"))
        #     logging.debug(f"Setting vertical alignment to {style.get('vertical_align')}")
        #     logging.debug(alignment)

        font = Font(name=style.get("font"), size=style.get("size"), bold=style.get("bold"), color=style.get("fgcolor"), strike=style.get("strike"))

        if style.get("bgcolor"):
            fill = PatternFill(start_color=style.get("bgcolor"), end_color=style.get("bgcolor"), fill_type="solid")
        else:
            fill = None

        # Bottom border white
        border = None
        if style.get("border"):
            border = Border(bottom=Side(style="thin", color="FFFFFF"))

        if not isinstance(item, tuple):
            item = [item]
        for cell in item:
            if alignment:
                cell.alignment = alignment 
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if border:
                cell.border = border



class HtmlWriter:
    def __init__(self, output_path):
        self.output_path = output_path
        self.output = open(output_path, "w")
        self.columns = COLUMNS
        self.output.write("""<!doctype html>
<html lang='en'>
<head>
    <meta content='text/html; charset=utf-8' http-equiv='Content-Type'>
    <meta name='viewport' content='width=device-width, initial-scale=1'>
    <title>PACT Simplified Data Model</title>
    <link rel='stylesheet' href='../assets/markdown.css'>
    <style>
    body       { padding: 0; margin: 0; }    
    table,th,tr,td { border-collapse: collapse; padding: 0; margin: 0; border: 0; vertical-align: top; }
    td,th      { padding-right: 4px; }
    h1,h2,h3   { border: none; margin: 0; }
    p          { margin: 4px; }
    li p       { margin: 0; }
    .title     { background-color: #08094C; color: #FFFFFF; font-size: 16px; }
    .subtitle  { background-color: #08094C; color: #FFFFFF; font-size: 16px; }
    .header    { background-color: #2A4879; color: #FFFFFF; }
    .subheader { background-color: #2A4879; color: #FFFFFF; }
    .subheader th { font-size: 11px; font-weight: normal; }
    .heading   { font-weight: bold; background-color: #489F81; color: #FFFFFF; }
    .bold      { font-weight: bold; }
    .small     { font-size: 0.9em; }
    .w15       { min-width: 100px; }
    .w17       { min-width: 120px; }
    .w20       { min-width: 140px; }
    .w30       { min-width: 210px; }
    .w60       { min-width: 420px; }
    </style>
</head>
<body>
    <table style='width:100%; border-collapse: collapse;'>
""")

    def save(self):
        self.output.write("</table>")
        self.output.write("</body></html>")
        self.output.close()

    def write_title(self, title):
        self.output.write(f"""
<tr class='title'><td colspan='{len(self.columns)}'><img src="../assets/logo-dark-margin.png" width="200" height="200"></td></tr>
<tr class='title'><td colspan='{len(self.columns)}'><h2>{title}</h2></td></tr>""")
 
    def write_subtitle(self, subtitle):
        self.output.write(f"<tr class='subtitle'><td colspan='{len(self.columns)}'><h2>{subtitle}</h2></td></tr>\n")

    def write_header(self):
        self.output.write("<tr class='header'>\n")
        for column in self.columns.values():
            self.output.write(f"<th class='w{column.get('width','')}'>{column.get('title','')}</th>")
        self.output.write("</tr>\n")
        self.output.write("<tr class='subheader'>\n")
        for name, column in self.columns.items():
            value = markdown.markdown(column.get("description","").replace("\n", "\n\n"))
            if name == "reporting":
                self.output.write(f"<th colspan='2'>{value}</th>")
            elif name != "description":
                self.output.write(f"<th>{value}</th>")
                
        self.output.write("</tr>\n")

    def write_type(self, items: list):
        self.output.write("<tr class='heading'>\n")
        for item in items:
            self.output.write(f"<td>{item}</td>")
        for i in range(len(self.columns) - len(items)):
            self.output.write("<td></td>")
        self.output.write("</tr>\n")

    def write_property(self, object):
        self.output.write("<tr>\n")
        
        # iterate over name and column in self.columns
        for name, column in self.columns.items():
            value = str(object.get(name, ""))
            style = column.get("style", "normal")
            if style == "text":
                value = markdown.markdown(value.replace("\n", "\n\n"))
            if style != "normal":
                css = f" class='{style}'"
            else:
                css = ""
            self.output.write(f"<td {css}>{value}</td>")

        self.output.write("</tr>\n")


def generate(writer, title, schema, typename):
    """
    @param ws: the worksheet to write to
    @param schema: the OpenAPI schema
    @param types: list of types to include in the worksheet
    
    This function generates an Excel worksheet from an OpenAPI schema. 
    It starts with the types specified in the list and follows all
    references to other types. 
    """
    
    # Validation rules legenda:
    validation_rules_legenda = schema["info"].get("x-rule","").strip()
    methodology_sections = schema["info"].get("x-methodology-sections", {})
    logging.info(methodology_sections)
    writer.columns["reporting"]["description"] = validation_rules_legenda

    # Append the title and header rows
    writer.write_title(title)
    writer.write_subtitle(schema["info"]["version"])

    # Append the column headers
    writer.write_header()

    # Inner function to get a succinct type description
    def get_type_description(info):
        type_description = ""
        if info.get("type", "") == "array":
            type_description = "array: " + get_type_description(info["items"])
        if info.get("type", "") == "object":
            if info.get("title"):
                type_description = info.get("title")
            else:
                type_description = "object"
        type_description += " " + info.get("format", "")
        if info.get("comment"):
            type_description += " (" + info["comment"] + ")\n"
        else:
            type_description += "\n"
        type_description += "|".join(info.get("enum", [])) + "\n"
        type_description = type_description.strip()
        if (type_description == ""):
            type_description = info.get("type","")
        elif not info.get("type","") in ["object","array"]:
            type_description += "\n(" + info.get("type","") + ")"
        return type_description

    # Inner function to write a property to the worksheet
    def write_property(name, info, parent, level):

        # Extract the type and description of the property
        type = info.get("type", "")
        
        if type == "object":
            write_type(name, info, level + 1)
            return
        
        type_description = get_type_description(info)
        description = info.get("summary") or info.get("description") or "N/A"
        
        for key, value in filter_text.items():
            description = description.replace(key, value)
        
        paragraphs = description.split("\n")
        description = ""
        for paragraph in paragraphs:
            if paragraph.strip() == "":
                description += "\n"
            elif paragraph.strip().startswith("-"):
                description += "\n" + paragraph.strip()
            else:
                description += paragraph.strip() + " "
        description = description.strip()
        logging.debug(description)
        
        # experiment: change the word property to attribute, use regex for word boundary
        description = re.sub(r'\bproperty\b', 'attribute', description)

        examples = info.get("examples", []) + ['','','']
        logging.debug(examples)
        mandatory = name in parent.get("required", [])
        
        # Append a row to the worksheet
        rule = info.get("x-rule", "")
        logging.info(f"Rule: {rule} mandatory: {mandatory} name: {name}")
        if mandatory and rule != "" and rule != "SHALL":
            raise Exception("Conflicting required status for " + name)
        if rule == "":
            if mandatory:
                rule = "SHALL" #"MUST"
            else:
                rule = "MAY"


        section = info.get("x-methodology", "")
        section = str(section) + " " + methodology_sections.get(section, "")
        section = section.strip()
        logging.info(section)

        writer.write_property(dict(
            property = name,
            attribute = info.get("x-term", ""),
            section = section,
            reporting = rule,
            comment = info.get("title", "") + info.get("x-comment", "") + info.get("note", ""),
            description = description,
            category = parent.get("title", ""),
            unit = info.get("x-unit", "-"),
            accepted = type_description,
            example1 = examples[0],
            example2 = examples[1],
            example3 = examples[2]
            ))

        if info.get("type", None) == "array" and info["items"].get("type") == "object":
            logging.debug(f"Writing array for {name} at level {level}")
            write_type(None, info["items"], level + 1)

        
    # Inner function to write a type to the worksheet
    def write_type(name, info, level=0):
        logging.debug(f"Writing type {name} at level {level}")
        if info.get("title") and name:
            # Append a row for the type itself and set background color to blue
            writer.write_type([name + ": " + info["title"], info.get("x-term"), info.get("x-methodology"), info.get("x-rule"), info.get("summary")])

        for prop_name, prop_info in info.get("properties", {}).items():
            # Skip obsolete properties
            if "obsolete" in prop_info:
                continue

            # Extract the type and description of the property
            logging.debug(f"Writing property {prop_name}")

            write_property(prop_name, prop_info, info, level)


    # Find the specified types in the schema
    type = schema["components"]["schemas"][typename]
    write_type(typename, type)




def openapi_to_excel(input_path, output_path:str, title, type):
    """
    @param input_path: path to the OpenAPI schema file
    @param output_path: path to the output Excel file
    @param title: title of the Excel worksheet
    @param types: name of the type to include in the Excel worksheet
    """

    # Load the schema from the file
    with open(input_path) as file:
        schema = yaml.safe_load(file)
    schema = jsonref.replace_refs(schema, merge_props=True)


    # Generate the Excel file
    writer = ExcelWriter(output_path)
    generate(writer, title, schema, type)
    writer.save()

    # Generate the HTML file
    output_path = output_path.removesuffix(".xlsx") + ".html"
    writer = HtmlWriter(output_path)
    generate(writer, title, schema, type)
    writer.save()



# if __name__ == "__main__":
#     # Get command line args
#     if len(sys.argv) < 2:
#         print("Usage: python3 generate-excel.py <input-path>")
#         print("This script generates an Excel file from a OpenAPI schema.")
#         print("")
#         print("Example:")
#         print("python3 generate-excel pact-openapi-2.2.1-wip.yaml")
#         print()
#         exit()
#     input_path = sys.argv[1]
#     if not os.path.exists(input_path):
#         print("File not found:", input_path)
#         exit()
#     status = " (Living Document)"
#     if (len(sys.argv) >= 3):
#         status = " (" + sys.argv[2].upper() + ")"
    
#     # Load the schema from the file
#     with open(input_path) as file:
#         schema1 = yaml.safe_load(file)
#     schema = jsonref.replace_refs(schema1, merge_props=True)

#     # Create a new workbook and select the active worksheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "PACT Simplified Data Model"
#     ws.sheet_view.zoomScale = 140

#     generate_excel(ws, schema, ["ProductFootprint"])

#     # Save the workbook to a file
#     output_path = os.path.basename(input_path)
#     output_path = output_path.replace('-openapi-', '-simplified-model-')
#     output_path = output_path.replace(".yaml", "") + ".xlsx"
#     output_path = os.path.join(os.path.dirname(input_path), output_path)
#     wb.save(output_path)
